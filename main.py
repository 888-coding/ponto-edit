# main.py
import os 
import openpyxl 
from datetime import time 
import colorama 
from colorama import Fore 

# 程式開始, 尋找 excel 檔案, 檔案位子
nome_arquivo = "file.xlsx"
diretorio_atual = os.path.dirname(os.path.realpath(__file__))

# 開啟檔案 , 開啟列表
wb = openpyxl.load_workbook(nome_arquivo)
todas_planilhas = wb.sheetnames
print(Fore.GREEN + "Bem vindo ao sistema" + Fore.RESET)

# 改日期
desejaAlterarData = input(Fore.YELLOW + "Deseja corrigir a o mes e ano ? (S/N)  : " + Fore.RESET).upper()
if desejaAlterarData == "S":
	novoMes = int(input("Favor digite o mês : "))
	novoAno = int(input("Favor digite o ano : "))
	sh = wb['DATA']
	sh['B1'].value = novoMes
	sh['B2'].value = novoAno	
	wb.save(nome_arquivo)
	print(Fore.GREEN + "Gravado com sucesso, abra de novo para alterar." + Fore.RESET)
	# 關閉程式
	exit()

#B2 插入時間(小時和分鐘) 
#sheet = arquivo['PEDRO']
#sheet['B2'].value = time(9,00,00)
#arquivo.save(nome_arquivo)

# 尋找名字 
sheetNome = input("Nome: ").upper()

nome_planilhas = wb.sheetnames

if sheetNome in nome_planilhas:
	# 找到了名字
	print(Fore.GREEN + f"Encontrado: {sheetNome}" + Fore.RESET)
	sh = wb[sheetNome]

	# 從開始還是那一天填寫?
	desejaContinuar = input("Deseja continuar? S/N : ").upper()
	if desejaContinuar == "S":
		diaInicio = int(input("Digite o dia de inicio : "))
	else:
		diaInicio = 1
	
	diaInicio = diaInicio + 1
	# Looping 的開啟 從第一天到最後一天
	for linha in range(diaInicio, 33):
		# 列表分為4行: B,C,D,E 
		# 分別為 上班時間, 吃飯時間, 吃飯回來時間, 下班時間
		# lista = ['B', 'C', 'D', 'E']
		campo = 'A' + str(linha)
		print(Fore.YELLOW + f"Dia : {sh[campo].value}" + Fore.RESET)
		print("**Entrada")
		# 進來時間 (小時) 
		while True: 
			horas_entrada = input("Horas : ")
			if horas_entrada.isnumeric():
				if int(horas_entrada) >= 0 and int(horas_entrada) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		# 進來時間 (分鐘) 
		while True:
			minutos_entrada = input("Minutos : ")
			if minutos_entrada.isnumeric():
				if int(minutos_entrada) >= 0 and int(minutos_entrada) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		print("**Almoço")
		# 吃飯時間 (小時)
		while True:
			horas_saida_almoco = input("Horas : ")
			if horas_saida_almoco.isnumeric():
				if int(horas_saida_almoco) >= 0 and int(horas_saida_almoco) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		# 吃飯時間 (分鐘) 
		while True:
			minutos_saida_almoco = input("Minutos : ")
			if minutos_saida_almoco.isnumeric():
				if int(minutos_saida_almoco) >= 0 and int(minutos_saida_almoco) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		print("**Volta almoço")
		# 吃飯回來時間 (小時)
		while True:
			horas_volta_almoco = input("Horas : ")
			if horas_volta_almoco.isnumeric():
				if int(horas_volta_almoco) >= 0 and int(horas_volta_almoco) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		# 吃飯回來時間 (分鐘) 
		while True:
			minutos_volta_almoco = input("Minutos : ")
			if minutos_volta_almoco.isnumeric():
				if int(minutos_volta_almoco) >= 0 and int(minutos_volta_almoco) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		print("**Saída")
		# 下班時間 (小時) 
		while True:
			horas_saida = input("Horas : ")
			if horas_saida.isnumeric():
				if int(horas_saida) >= 0 and int(horas_saida) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)
		# 下班時間 (分鐘)
		while True:
			minutos_saida = input("Minutos : ")
			if minutos_saida.isnumeric():
				if int(minutos_saida) >= 0 and int(minutos_saida) <= 100:
					break
				else:
					print(Fore.RED + "Erro!" + Fore.RESET)
			else:
				print(Fore.RED + "Erro!" + Fore.RESET)

		# 顯示所寫的時間
		print(Fore.GREEN + f"Voce digitou: [ENTRADA: {horas_entrada}:{minutos_entrada}], [ALMOÇO: {horas_saida_almoco}:{minutos_saida_almoco}], [VOLTA ALMOCO: {horas_volta_almoco}:{minutos_volta_almoco}], [SAÍDA: {horas_saida}:{minutos_saida}]" + Fore.RESET)

		# 如果想要save請按 '1'
		while True:
			desejaSalvar = input("Deseja salvar o horário? (Sim 1/Nao 0) ").upper()
			if desejaSalvar == "1" or desejaSalvar == "0":
				break
			else:
				print("Escolha Sim 1 ou Nao 0")
		if desejaSalvar == "1":
			campo = 'B' + str(linha)
			sh[campo].value = time(int(horas_entrada),int(minutos_entrada),00)
			campo = 'C' + str(linha)
			sh[campo].value = time(int(horas_saida_almoco),int(minutos_saida_almoco),00)
			campo = 'D' + str(linha)
			sh[campo].value = time(int(horas_volta_almoco),int(minutos_volta_almoco),00)
			campo = 'E' + str(linha)
			sh[campo].value = time(int(horas_saida),int(minutos_saida),00)
			wb.save(nome_arquivo)
			print(Fore.GREEN + "Gravado com sucesso" + Fore.RESET)

else:
	#沒有找到名字
	print(Fore.RED + "Nome errado" + Fore.RESET)